#教育机构 ：马士兵教育
#讲    师：杨淑娟
import  openpyxl

workbook=openpyxl.load_workbook('新表.xlsx')
sheet=workbook['Sheet1']
#拆分合并单元格
sheet.unmerge_cells('D1:G2')
sheet.unmerge_cells(start_row=1,start_column=8,end_row=5,end_column=12)
workbook.save('新表.xlsx')