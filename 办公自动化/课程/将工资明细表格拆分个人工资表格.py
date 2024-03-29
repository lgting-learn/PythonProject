# 教育机构 ：马士兵教育
# 讲    师：杨淑娟
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border

# 将工资明细表格拆分个人工资表格
# 设置字体样式
font = Font(name='宋体', size=20, bold=True)
font2 = Font(name='宋体', size=12, bold=True)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
side = Side(style='thin', color='000000')
border = Border(left=side, top=side, right=side, bottom=side)

# 加载Excel文件
workbook = openpyxl.load_workbook('工资明细.xlsx')
sheet = workbook['Sheet1']

# 获取指定的单元格的值
al_value = sheet['A1'].value
l2_value = sheet['L2'].value
# print(al_value)

# 获取指定的行
rows = sheet[3]
lst_value = []
for cell in rows:
    lst_value.append(cell.value)
print('lst_value==',lst_value)
# print(lst_value)
for i in range(4, 32):
    # 将表头部分写进新的Excel文件中
    write_workbook = openpyxl.Workbook()
    write_sheet = write_workbook.active

    # 第一行为合并单元格
    write_sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=12)
    write_sheet.row_dimensions[1].height = 25.8
    write_sheet['A1'] = al_value
    write_sheet['A1'].font = font
    write_sheet['A1'].alignment = alignment

    # 发放薪水的单位
    write_sheet['L2'] = l2_value
    write_sheet['L2'].font = font2
    # 第三行
    write_sheet.append(lst_value)
    three_rows = write_sheet[3]
    for cell in three_rows:
        cell.font = font2
        cell.alignment = alignment
        cell.border = border
    write_sheet.row_dimensions[3].height = 40.8
    # 设置列宽
    thr_rows = write_sheet[3]
    for col in thr_rows:
        write_sheet.column_dimensions[col.column_letter].width = 13

    # 向Excel表格中添加数据
    write_sheet['A4'].value = sheet['A' + str(i)].value
    write_sheet['B4'].value = sheet['B' + str(i)].value
    write_sheet['C4'].value = sheet['C' + str(i)].value
    write_sheet['D4'].value = sheet['D' + str(i)].value
    write_sheet['E4'].value = sheet['E' + str(i)].value
    write_sheet['F4'].value = '=SUM(C4:E4)'
    write_sheet['G4'].value = sheet['G' + str(i)].value
    write_sheet['H4'].value = sheet['H' + str(i)].value
    write_sheet['I4'].value = sheet['I' + str(i)].value
    write_sheet['J4'].value = sheet['J' + str(i)].value
    write_sheet['K4'].value = sheet['K' + str(i)].value
    write_sheet['L4'].value = '=F4-G4-H4-I4-J4-K4'

    for_row = write_sheet[4]
    for cell in for_row:
        cell.alignment = alignment
        cell.border = border
    write_workbook.save('salary/' + write_sheet['B4'].value + '.xlsx')
