# 教育机构 ：马士兵教育
# 讲    师：杨淑娟
import openpyxl
from openpyxl.styles import Font, PatternFill
import re

workbook = openpyxl.load_workbook('京东鞋子评论信息.xlsx')
# print(workbook.sheetnames)
# 标识关键字所在行
font = Font(color='ff0000')
fill = PatternFill(fill_type='solid', fgColor='ffff00')
sheet = workbook['mysheet']
workbook.copy_worksheet(sheet)
sheet.title = '好评数据0301'
for row in workbook['好评数据'].rows:
    if re.search('.*好.*', row[0].value):
        for i in range(0, 4):
            row[i].fill = fill
            row[i].font = font
workbook.save('京东鞋子评论信息.xlsx')
