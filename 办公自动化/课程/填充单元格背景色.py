# 教育机构 ：马士兵教育
# 讲    师：杨淑娟
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import GradientFill

workbook = openpyxl.load_workbook('新表.xlsx')
sheet = workbook['Sheet1']
cell_c5 = sheet['C5']
# 填充单元格背景色
pattern_fill = PatternFill(fill_type='solid', fgColor='ffff00')
cell_c5.fill = pattern_fill
workbook.save('新表.xlsx')
