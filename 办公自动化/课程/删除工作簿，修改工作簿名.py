#教育机构 ：马士兵教育
#讲    师：杨淑娟

import openpyxl
workbook=openpyxl.load_workbook('京东鞋子评论信息.xlsx')
print(workbook.sheetnames)
#workbook.create_sheet('mysheet')
#删除工作簿
#sheet=workbook['mysheet']
#workbook.remove(sheet)
#修改工作簿名
sheet=workbook['评论信息1']
#workbook.copy_worksheet(sheet)
sheet.title='评论信息'

print(workbook.sheetnames)

workbook.save('京东鞋子评论信息.xlsx')