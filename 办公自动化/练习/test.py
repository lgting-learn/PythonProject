import openpyxl
#将工资明细表格拆分个人工资表格
workbook = openpyxl.load_workbook('工资明细.xlsx')
sheet = workbook['Sheet1']
fir_value = sheet['A1'].value
sec_value = sheet['L2'].value
third_rows = sheet[3]
third_arr = []
for i in third_rows:
    third_arr.append(i.value)
rows = sheet[4:31]
lst = []
for row in rows:
    lit = []
    for i in range(0, 12):
        lit.append(row[i].value)
    lst.append(lit)

for i in range(4,32):
    new_book = openpyxl.Workbook()
    new_sheet = new_book.active
    #!!!合并单元格
    new_sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=12)
    new_sheet['A1'] = fir_value
    new_sheet['L2'] = sec_value
    print(third_arr)
    #third_arr是普通数组的话，new_sheet可直接追加
    new_sheet.append(third_arr)
    new_sheet['A4'] = sheet['A'+str(i)].value
    new_sheet['B4'] = sheet['B'+str(i)].value
    new_sheet['C4'] = sheet['C'+str(i)].value
    new_sheet['D4'] = sheet['D'+str(i)].value
    new_sheet['E4'] = sheet['E'+str(i)].value
    #!!!注意
    new_sheet['F4'] = '=SUM(C4:E4)'
    new_sheet['G4'] = sheet['G'+str(i)].value
    new_sheet['H4'] = sheet['H'+str(i)].value
    new_sheet['I4'] = sheet['I'+str(i)].value
    new_sheet['J4'] = sheet['J'+str(i)].value
    new_sheet['K4'] = sheet['K'+str(i)].value
    new_sheet['L4'] = '=F4-G4-H4-I4-J4-K4'
    print(sheet['L'+str(i)].value)
    new_book.save('AAA/'+sheet['B'+str(i)].value+'.xlsx')
