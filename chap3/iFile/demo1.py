import openpyxl

file_init = '原表.xlsx'
file_map = '映射表.xlsx'
workbookInit = openpyxl.load_workbook(file_init)
workbookMap = openpyxl.load_workbook(file_map)
sheetInit = workbookInit['Sheet1']
sheetMap = workbookMap['Sheet1']
rowsInitA = sheetInit['A']
rowsInitB = sheetInit['B']
rowsMapA = sheetMap['A']
rowsMapB = sheetMap['B']

# 修改单元格的值
def setCelValue(row, column, cellValue):
    sheetInit.cell(row=row, column=column, value=cellValue)

if __name__ == '__main__':
    # 修改单元格
    for rowInit in rowsInitA:
        for rowMap in rowsMapA:
            if rowInit.value == rowMap.value:
                setCelValue(rowInit.row, 1, rowsMapB[rowMap.row - 1].value)
    # 在地区部前添加0
    for row in rowsInitB:
        if (row.value != '地区部'):
            setCelValue(row.row, 2, '0' + str(row.value))
    workbookInit.save(file_init)
