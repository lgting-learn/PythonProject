import openpyxl
workbook = openpyxl.load_workbook('京东鞋子评论信息.xlsx')
print(workbook.sheetnames)
#插入sheet
# workbook.create_sheet('lgting')
#删除sheet
# sheet=workbook['lgting']#获取sheet
# workbook.remove(sheet)
# 复制sheet
sheet=workbook['评论信息']
# workbook.copy_worksheet(sheet)
# 重命名sheet
# sheet.title='评论信息 test'
# 冻结窗格
# sheet.freeze_panes='A2'#冻结第1行标题
#选中筛选
sheet.auto_filter.ref = sheet.dimensions
workbook.save('京东鞋子评论信息.xlsx')
